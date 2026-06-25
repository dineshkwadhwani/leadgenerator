"""
Tracksoft Lead Processor — Flask web application.
Run on port 5001: python app.py
"""

import logging
import os
import secrets
import time
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv
from flask import Flask, jsonify, render_template, request, send_file
from werkzeug.utils import secure_filename

load_dotenv(Path(__file__).parent / ".env")

import db
import email_helper
import worker

# ── App setup ──────────────────────────────────────────────────────────────────
BASE = Path(__file__).parent
UPLOAD_DIR = BASE / "uploads"
OUTPUT_DIR = BASE / "outputs"
TEMPLATE_FILE = BASE.parent / "Template.xlsx"
MAX_ROWS = 100
MAX_FILE_MB = 20

UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = MAX_FILE_MB * 1024 * 1024

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)
logger = logging.getLogger(__name__)

db.init_db()

# ── Column resolution (mirrors selenium_scraper logic) ─────────────────────────
INPUT_COL_ALIASES = {
    "location": ["location", "city", "city_name", "taluka_name"],
    "org_name": ["organization name", "org name", "school_name", "name"],
    "org_type": ["org type", "org_type", "type"],
}


def _resolve(headers_lower: list, aliases: list) -> bool:
    return any(a in headers_lower for a in aliases)


def _validate_sheet(df: pd.DataFrame):
    lowered = [c.lower().strip() for c in df.columns]
    missing = []
    for field, aliases in INPUT_COL_ALIASES.items():
        if not _resolve(lowered, aliases):
            missing.append(field.replace("_", " ").title())
    if missing:
        raise ValueError(
            f"Template columns not found: {', '.join(missing)}. "
            "Please use the official template."
        )


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/template")
def download_template():
    if not TEMPLATE_FILE.exists():
        return "Template file not found on server.", 404

    wb = load_workbook(str(TEMPLATE_FILE))
    ws = wb.active
    ws.insert_rows(1)
    ws["A1"] = "Please fill only first three columns."
    ws["A1"].font = ws["A2"].font.copy(bold=True)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="Tracksoft_Lead_Template.xlsx",
        mimetype=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
    )


@app.route("/api/upload", methods=["POST"])
def api_upload():
    f = request.files.get("file")
    if not f or f.filename == "":
        return jsonify(error="No file uploaded."), 400

    fname = secure_filename(f.filename)
    if not fname.lower().endswith(".xlsx"):
        return jsonify(error="Only .xlsx files are supported."), 400

    try:
        df = pd.read_excel(f)
    except Exception:
        return jsonify(error="Could not read the file. Make sure it is a valid .xlsx."), 400

    if len(df) == 0:
        return jsonify(error="The file has no data rows."), 400

    if len(df) > MAX_ROWS:
        return jsonify(
            error=(
                f"Only {MAX_ROWS} rows at a time can be processed. "
                "For larger files please send the document to "
                "contact@tracksoftsolutions.com"
            )
        ), 400

    try:
        _validate_sheet(df)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400

    job_id = secrets.token_hex(16)
    input_path = str(UPLOAD_DIR / f"{job_id}_input.xlsx")
    output_name = fname.replace(".xlsx", "_LEADS.xlsx")
    output_path = str(OUTPUT_DIR / f"{job_id}_output.xlsx")

    df.to_excel(input_path, index=False, engine="openpyxl")
    db.create_job(job_id, input_path, output_path, output_name, len(df))
    worker.start(job_id, input_path, output_path)

    logger.info("Job %s started — %s rows", job_id, len(df))
    return jsonify(jobId=job_id, rowCount=len(df))


@app.route("/api/status/<job_id>")
def api_status(job_id: str):
    job = db.get_job(job_id)
    if not job:
        return jsonify(error="Job not found."), 404
    return jsonify(
        status=job["status"],
        message=job["message"] or "",
        processed=job["processed"] or 0,
        rowCount=job["row_count"] or 0,
    )


@app.route("/api/request-otp", methods=["POST"])
def api_request_otp():
    body = request.get_json(silent=True) or {}
    job_id = str(body.get("jobId", "")).strip()
    email = str(body.get("email", "")).strip().lower()
    mobile = str(body.get("mobile", "")).strip()

    if not job_id or not email or not mobile:
        return jsonify(error="jobId, email and mobile are required."), 400

    if "@" not in email or "." not in email.split("@")[-1]:
        return jsonify(error="Please enter a valid email address."), 400

    if len(mobile) < 10:
        return jsonify(error="Please enter a valid mobile number."), 400

    job = db.get_job(job_id)
    if not job:
        return jsonify(error="Job not found."), 404
    if job["status"] != "complete":
        return jsonify(error="File is not ready yet. Please wait."), 400

    otp = str(secrets.randbelow(900000) + 100000)   # 6-digit
    db.save_otp(job_id, otp, email, mobile)

    try:
        email_helper.send_otp(email, otp)
    except Exception as exc:
        logger.error("OTP send failed: %s", exc)
        return jsonify(error=f"Could not send OTP email: {exc}"), 500

    logger.info("OTP sent for job %s to %s", job_id, email)
    return jsonify(message="OTP sent to your email.")


@app.route("/api/verify-otp", methods=["POST"])
def api_verify_otp():
    body = request.get_json(silent=True) or {}
    job_id = str(body.get("jobId", "")).strip()
    email = str(body.get("email", "")).strip().lower()
    otp_in = str(body.get("otp", "")).strip()

    if not job_id or not email or not otp_in:
        return jsonify(error="jobId, email and otp are required."), 400

    otp_state = db.get_otp(job_id)
    if not otp_state:
        return jsonify(error="OTP expired or not found. Please request again."), 400

    if otp_state["email"] != email:
        return jsonify(error="Email does not match OTP request."), 400

    if int(time.time()) > otp_state["expires_at"]:
        db.delete_otp(job_id)
        return jsonify(error="OTP has expired. Please request a new one."), 400

    if otp_state["attempts"] >= 3:
        return jsonify(
            error="Too many incorrect attempts. Please request a new OTP."
        ), 400

    if otp_state["otp"] != otp_in:
        db.increment_otp_attempts(job_id)
        remaining = 3 - (otp_state["attempts"] + 1)
        return jsonify(
            error=f"Incorrect OTP. {remaining} attempt(s) remaining."
        ), 400

    job = db.get_job(job_id)
    if not job or not Path(job["output_path"]).exists():
        return jsonify(
            error="Output file not found. Please upload your file again."
        ), 400

    try:
        file_bytes = Path(job["output_path"]).read_bytes()
        email_helper.send_result(email, job["output_name"], file_bytes)
    except Exception as exc:
        logger.error("Result email failed: %s", exc)
        return jsonify(error=f"Could not send file: {exc}"), 500

    # Clean up
    db.delete_otp(job_id)
    try:
        Path(job["input_path"]).unlink(missing_ok=True)
        Path(job["output_path"]).unlink(missing_ok=True)
    except Exception:
        pass

    logger.info("Result emailed for job %s to %s", job_id, email)
    return jsonify(message="Verified! Your lead file has been sent to your email.")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    app.run(host="0.0.0.0", port=port, debug=False)
